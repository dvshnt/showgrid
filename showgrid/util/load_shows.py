# Upload .xls file
# Read shows and place them by venue

import re
import sys
import time
from datetime import datetime
import openpyxl as px

from django.core.management import setup_environ

import sys
sys.path.append('/home/ubuntu/app/showgrid/')
from showgrid import settings
setup_environ(settings)

from showgrid.models import Venue, Show
from django.core.files import File

def get_time(date, hour):
	if ":" in hour:
		return time.strptime(date + " " + hour, "%m/%d/%Y %I:%M%p")
	return time.strptime(date + " " + hour, "%m/%d/%Y %I%p")

def get_venue(name):
	if "Lindsley" in name:
		return Venue.objects.get(name="3rd & Lindsley")
	elif "Porter" in name:
		return Venue.objects.get(name="12th & Porter")
	elif "Exit" in name:
		return Venue.objects.get(name="Exit/In")
	elif "Marathon" in name:
		return Venue.objects.get(name="Marathon Music Works")
	elif "Basement" in name:
		return Venue.objects.get(name="The Basement")
	elif "Rutledge" in name:
		return Venue.objects.get(name="The Rutledge")
	elif "Cannery" in name:
		return Venue.objects.get(name="The Cannery Ballroom")
	elif "Mercy" in name:
		return Venue.objects.get(name="Mercy Lounge")
	elif "Watt" in name:
		return Venue.objects.get(name="The High Watt")
	elif "Bluebird" in name:
		return Venue.objects.get(name="Bluebird Cafe")
	elif "End" in name:
		return Venue.objects.get(name="The End")
	elif "Listening" in name:
		return Venue.objects.get(name="The Listening Room")
	elif "Hard" in name:
		return Venue.objects.get(name="Hard Rock Cafe")
	elif "Douglas" in name:
		return Venue.objects.get(name="Douglas Corner")
	elif "Wildhorse" in name:
		return Venue.objects.get(name="Wildhorse Saloon")
	elif "Opry" in name:
		return Venue.objects.get(name="The Opry")
	elif "Ryman" in name:
		return Venue.objects.get(name="Ryman Auditorium")
	elif "Bridgestone" in name:
		return Venue.objects.get(name="Bridgestone Arena")
	elif "Belcourt" in name:
		return Venue.objects.get(name="Belcourt Taps")
	elif "Pucketts" in name:
		return Venue.objects.get(name="Pucketts")
	elif "Loveless" in name:
		return Venue.objects.get(name="Music City Roots (Loveless Cafe)")
	elif "Stone" in name:
		return Venue.objects.get(name="The Stone Fox")
	elif "Zanie" in name:
		return Venue.objects.get(name="Zanie\'s")
	elif "Station" in name:
		return Venue.objects.get(name="Station Inn")
	else:
		return Venue.objects.get(name=name)

def read_shows(wb):
	for sheet in wb.worksheets:
		for i in range(2,9):
			date = sheet.cell(row=2, column=i).value
			date = re.split(r'\s', date)[1] + '/2014'
			print date
			for j in range(3,27):
				artist = []
				hour = []
				venue = get_venue(sheet.cell(row=j, column=1).value)

				content = sheet.cell(row=j, column=i)
				
				try:
					shows = re.split(r'\s/\s', content.value.encode('ascii', 'ignore'))
					for show in shows:
						info = re.split(r'\s', show)
						for part in info:
							if not "PM" in part and not "AM" in part:
								artist.append(part)
							else:
								hour.append(part)

						artist = ' '.join(artist) 
						hour = ''.join(hour)

						date_time = datetime.fromtimestamp(time.mktime(get_time(date, hour)))

						show_entry = Show(band_name=artist, venue=venue, date=date_time)
						show_entry.save()
				except:
					continue


read_shows(px.load_workbook(filename=sys.argv[1]))
